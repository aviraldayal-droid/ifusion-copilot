"""
TBG Excel file parser.

Column layout varies by month: later months have ACTU1/2/3 columns inserted
between Budget and N-1. We read the actual column-header row to build the
value-type → column-index map for each month, instead of assuming a fixed offset.

Metric code pattern: PL1, Opex3, CA10, PARC7, CAP4 (letters + digits).
"""
from __future__ import annotations

import re
from datetime import datetime
from typing import Any

import openpyxl

SHEETS_OF_INTEREST: dict[str, str] = {
    "P&L conso": "pnl_conso",
    "CA Mobile": "ca_mobile",
    "Opex Consolidés": "opex_consolides",
    "Capex Consolidés": "capex_consolides",
    "Mobile Money": "mobile_money",
    "Parc Mobile ": "parc_mobile",
    "Marge Mobile": "marge_mobile",
    "Trafic mobile": "trafic_mobile",
    "Data Mobile": "data_mobile",
    "Cash conso": "cash_conso",
    "Marge Fixe": "marge_fixe",
    "CA Fixe": "ca_fixe",
}

_CODE_RE = re.compile(r"^[A-Z][A-Za-z]{1,10}\d+$")

_SKIP_LABELS = frozenset(
    {"0", "None", "(En Millions ML)", "En monnaie locale", ""}
)


def _is_numeric(value: Any) -> bool:
    return isinstance(value, (int, float)) and (value == value)


def _clean(raw: Any) -> str:
    if raw is None:
        return ""
    return re.sub(r"\s+", " ", str(raw)).strip()


def _find_date_row(rows: list[tuple]) -> tuple[int, dict[int, str]]:
    """
    Return (row_index, {col_index: 'YYYY-MM'}) for the first row with ≥3 datetimes.
    Returns (-1, {}) if not found.
    """
    for i, row in enumerate(rows[:20]):
        col_dates: dict[int, str] = {}
        for col_idx, cell in enumerate(row):
            if isinstance(cell, datetime):
                col_dates[col_idx] = cell.strftime("%Y-%m")
        if len(col_dates) >= 3:
            return i, col_dates
    return -1, {}


def _build_col_map_from_headers(
    date_cols: dict[int, str],
    header_row: tuple,
) -> dict[str, dict[str, int]]:
    """
    For each month, scan the header row from date_col until the next date_col
    and map recognised column types to their column indices.

    Recognised types  (priority order within a month block):
      reel          – first "Réel" occurrence
      budget        – "Budget" (not preceded by "Ecart")
      ecart_budget  – "Ecart" + "Budget"
      n1_reel       – second "Réel" occurrence  OR  "N-1" / "2024" / "2023"
      evol_pct      – "Evol" / "%"
    Ignored:
      Actu columns and Ecart/Actu columns.
    """
    sorted_date_cols = sorted(date_cols.keys())
    total_cols = len(header_row) if header_row else 200

    col_map: dict[str, dict[str, int]] = {}

    for i, date_col in enumerate(sorted_date_cols):
        period = date_cols[date_col]
        next_date_col = sorted_date_cols[i + 1] if i + 1 < len(sorted_date_cols) else total_cols

        vtype_cols: dict[str, int] = {}
        reel_count = 0

        for col_idx in range(date_col, min(next_date_col, total_cols)):
            if col_idx >= len(header_row):
                break
            h = _clean(header_row[col_idx]).lower()
            if not h:
                continue

            # Skip ACTU and Ecart/ACTU
            if "actu" in h:
                continue
            # Ecart vs Budget (keep); Ecart vs Actu (skip)
            if "ecart" in h and "budget" not in h:
                continue

            if "réel" in h or "reel" in h:
                reel_count += 1
                if reel_count == 1:
                    vtype_cols["reel"] = col_idx
                else:
                    vtype_cols.setdefault("n1_reel", col_idx)
            elif "budget" in h and "ecart" not in h:
                vtype_cols.setdefault("budget", col_idx)
            elif "ecart" in h and "budget" in h:
                vtype_cols.setdefault("ecart_budget", col_idx)
            elif "evol" in h or ("%" in h and "ecart" not in h):
                vtype_cols.setdefault("evol_pct", col_idx)

        col_map[period] = vtype_cols

    return col_map


def _extract_label_code(row: tuple, min_date_col: int) -> tuple[str, str | None]:
    """
    Scan all columns before min_date_col.
    Metric codes match _CODE_RE (e.g. PL1, Opex3).
    Everything else non-empty → label.
    """
    label = ""
    code: str | None = None
    for col_idx in range(min(min_date_col, len(row))):
        val = _clean(row[col_idx])
        if not val or val in _SKIP_LABELS:
            continue
        if _CODE_RE.match(val):
            code = val
        elif not label:
            label = val
    return label, code


def _extract_value(row: tuple, col_idx: int) -> float | None:
    if col_idx >= len(row):
        return None
    val = row[col_idx]
    if _is_numeric(val):
        return round(float(val), 4)
    return None


def _parse_sheet(ws) -> dict:
    all_rows = list(ws.iter_rows(values_only=True))

    date_row_idx, date_cols = _find_date_row(all_rows)
    if date_row_idx < 0 or not date_cols:
        return {}

    min_date_col = min(date_cols.keys())

    # Find header row: first non-empty row after date_row_idx that has text
    # in a column adjacent to a date column (not another date row).
    header_row: tuple = ()
    for candidate in all_rows[date_row_idx + 1 : date_row_idx + 5]:
        if candidate and _clean(candidate[min_date_col]).strip():
            header_row = candidate
            break
    if not header_row:
        # Fallback: assume standard 5-col layout
        header_row = tuple()

    col_map = _build_col_map_from_headers(date_cols, header_row)
    sorted_periods = sorted(col_map.keys())

    metrics: dict[str, dict] = {}
    current_section: str = ""

    for row in all_rows[date_row_idx + 2 :]:
        if not row or len(row) <= min_date_col:
            continue

        label, code = _extract_label_code(row, min_date_col)
        if not label or label in _SKIP_LABELS:
            continue

        metric_key = code if code else label

        values: dict[str, dict[str, float | None]] = {}
        for period, cols in col_map.items():
            if not cols:
                continue
            pv: dict[str, float | None] = {}
            for vtype, col_idx in cols.items():
                pv[vtype] = _extract_value(row, col_idx)
            if any(v is not None for v in pv.values()):
                values[period] = pv

        if not values:
            # Row has a label but no numeric data — treat as sub-section header
            current_section = label
            continue

        metrics[metric_key] = {"label": label, "code": code, "section": current_section, "values": values}

    return {"periods": sorted_periods, "metrics": metrics}


def parse_tbg_file(file_path: str) -> dict:
    """
    Parse a TBG Excel file. Returns:
      { "file": "...", "all_periods": [...], "sheets": { "pnl_conso": {...}, ... } }
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    available = set(wb.sheetnames)

    result: dict = {"file": file_path.split("/")[-1], "sheets": {}}

    for sheet_name, sheet_key in SHEETS_OF_INTEREST.items():
        if sheet_name not in available:
            continue
        ws = wb[sheet_name]
        parsed = _parse_sheet(ws)
        if parsed:
            result["sheets"][sheet_key] = parsed

    wb.close()

    all_periods: set[str] = set()
    for sd in result["sheets"].values():
        all_periods.update(sd.get("periods", []))
    result["all_periods"] = sorted(all_periods)

    return result


# ---------------------------------------------------------------------------
# Lookup helpers used by agent tools
# ---------------------------------------------------------------------------

def find_metric_by_label(
    parsed_data: dict, sheet_key: str, partial_label: str
) -> list[tuple[str, str]]:
    sheet = parsed_data["sheets"].get(sheet_key)
    if not sheet:
        return []
    lower = partial_label.lower()
    return [
        (key, m["label"])
        for key, m in sheet["metrics"].items()
        if lower in m["label"].lower() or lower in key.lower()
    ]


def compute_yoy_pct(current: float | None, prior: float | None) -> float | None:
    if current is None or prior is None or prior == 0:
        return None
    return round((current - prior) / abs(prior) * 100, 2)


def compute_vs_budget_pct(reel: float | None, budget: float | None) -> float | None:
    if reel is None or budget is None or budget == 0:
        return None
    return round((reel - budget) / abs(budget) * 100, 2)
